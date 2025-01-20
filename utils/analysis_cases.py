from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import json

def highlight_cells_with_string(filename, sheetname, columns, target_string, fill_color):
    # 加载Excel文件
    workbook = load_workbook(filename)
    # 选择工作表
    sheet = workbook[sheetname]

    # 遍历指定列的所有单元格
    for column in columns:
        for cell in sheet[column]:
            if cell.value and target_string in str(cell.value):
                # 创建填充样式对象
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                # 应用填充样式到单元格
                cell.fill = fill

    # 保存修改后的Excel文件
    workbook.save(filename)

if __name__ == '__main__':

    # data1 = pd.read_csv("unify_outputs/initial_queries_three_model_predictions.csv")
    # data1.to_excel("unify_outputs/initial_queries_three_model_predictions.xlsx", index=False, header=True)

    RES_DIR = "victim_responses"
    EVAL_DIR = "guard_eval_results/llama"
    TO_DIR = "unify_outputs"
    base_file_name = "llama-2-7b-mi-suffix-response-0508"
    chat_file_name = "llama-2-7b-chat-mi-suffix-response-0508"
    csv_name = "mi_three_model_predictions"

    llama_chat_pred = json.load(open(f"{RES_DIR}/{chat_file_name}.json"))
    llama_chat_pred_eval = json.load(open(f"{EVAL_DIR}/{chat_file_name}-guard-eval.json"))
    llama_base_pred = json.load(open(f"{RES_DIR}/{base_file_name}.json"))
    llama_base_eval = json.load(open(f"{EVAL_DIR}/{base_file_name}-guard-eval.json"))
    # vicuna_pred = json.load(open(f"{RES_DIR}/vicuna-7b-v1.5-response-0425.json"))
    # vicuna_pred_eval = json.load(open(f"{EVAL_DIR}/vicuna-7b-v1.5-response-0425-guard-eval.json"))

    # llama_chat_pred = json.load(open("victim_responses/llama-2-7b-chat-mi-response-0503.json"))
    # llama_chat_pred_eval = json.load(open("guard_eval_results/llama-2-7b-chat-mi-response-0503-guard-eval.json"))
    # llama_base_pred = json.load(open("victim_responses/llama-2-7b-mi-response-0503.json"))
    # llama_base_eval = json.load(open("guard_eval_results/llama-2-7b-mi-response-0503-guard-eval.json"))
    # vicuna_pred = json.load(open("victim_responses/vicuna-7b-v1.5-mi-response-0503.json"))
    # vicuna_pred_eval = json.load(open("guard_eval_results/vicuna-7b-v1.5-mi-response-0503-guard-eval.json"))

    prompts = []
    llama_chat_responses = []
    llama_base_responses = []
    vicuna_response = []

    llama_chat_evals = []
    llama_base_evals = []
    vicuna_evals = []
    for idx, item in enumerate(llama_chat_pred):
        prompts.append(item["prompt"])
        llama_chat_responses.append(item["model_prediction"])
        llama_base_responses.append(llama_base_pred[idx]["model_prediction"])
        # vicuna_response.append(vicuna_pred[idx]["model_prediction"])

        llama_chat_evals.append(llama_chat_pred_eval[idx]["model_prediction"])
        llama_base_evals.append(llama_base_eval[idx]["model_prediction"])
        # vicuna_evals.append(vicuna_pred_eval[idx]["model_prediction"])

    data = {
        "prompt": prompts,
        "llama_base": llama_base_responses,
        "llama_base_evals": llama_base_evals,
        # "vicuna": vicuna_response,
        # "vicuna_evals": vicuna_evals,
        "llama_chat": llama_chat_responses,
        "llama_chat_evals": llama_chat_evals,
    }
    pd_data = pd.DataFrame(data=data, columns=list(data.keys()))
    # pd_data.to_csv(f'{TO_DIR}/initial_queries_three_model_predictions.csv', index=False)  # 306
    # pd_data.to_csv(f'{TO_DIR}/{csv_name}.csv', index=False)  # 306
    # data = pd.read_csv(f"{TO_DIR}/{csv_name}.csv")
    pd_data.to_excel(f"{TO_DIR}/{csv_name}.xlsx", index=False, header=True)

    # 示例用法
    columns = ["C", "E"]  # 指定列 , "G"
    highlight_cells_with_string(f"{TO_DIR}/{csv_name}.xlsx",
                                "Sheet1",
                                columns,
                                "unsafe",
                                "FFFF00")